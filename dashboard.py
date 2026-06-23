"""Web dashboard for slip_ledger — login + transaction table, presentable for management.

Run:
  pip install -r requirements.txt
  fill in DATABASE_URL / DASHBOARD_SESSION_SECRET in .env
  uvicorn dashboard:app --host 0.0.0.0 --port 8081
"""

import io
import json
import os
import uuid
from collections import defaultdict
from datetime import date, time
from pathlib import Path

import bcrypt
import pandas as pd
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from linebot.v3.messaging import ApiClient, Configuration, MessagingApi
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from starlette.middleware.sessions import SessionMiddleware

from categorize import get_categories, guess_category

load_dotenv()

LINE_FILES_DIR = Path("line_files")
LINE_FILES_DIR.mkdir(exist_ok=True)

DATABASE_URL = os.environ["DATABASE_URL"]
SESSION_SECRET = os.environ["DASHBOARD_SESSION_SECRET"]
LINE_ACCESS_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")

THAI_MONTHS = {
    1: "ม.ค.", 2: "ก.พ.", 3: "มี.ค.", 4: "เม.ย.", 5: "พ.ค.", 6: "มิ.ย.",
    7: "ก.ค.", 8: "ส.ค.", 9: "ก.ย.", 10: "ต.ค.", 11: "พ.ย.", 12: "ธ.ค.",
}

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET, same_site="lax")
templates = Jinja2Templates(directory="templates")


def db():
    return psycopg2.connect(DATABASE_URL)


def require_login(request: Request):
    if not request.session.get("user"):
        return RedirectResponse("/login")
    return None


def require_admin(request: Request):
    redirect = require_login(request)
    if redirect:
        return redirect
    if request.session.get("role") != "admin":
        raise HTTPException(403, "ต้องเป็น admin ถึงแก้ไขได้")
    return None


@app.get("/login")
def login_form(request: Request):
    if request.session.get("user"):
        return RedirectResponse("/")
    return templates.TemplateResponse(request, "login.html", {"error": None})


@app.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    conn = db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT password_hash, role FROM dashboard_users WHERE username = %s", (username,))
            row = cur.fetchone()
    finally:
        conn.close()

    if not row or not bcrypt.checkpw(password.encode(), row[0].encode()):
        return templates.TemplateResponse(
            request, "login.html", {"error": "username/password ไม่ถูกต้อง"}, status_code=401
        )

    request.session["user"] = username
    request.session["role"] = row[1]
    return RedirectResponse("/", status_code=302)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login")


@app.get("/files/{raw_file_id}")
def serve_file(raw_file_id: int, request: Request):
    redirect = require_login(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT storage_path FROM raw_files WHERE id = %s", (raw_file_id,))
            row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        raise HTTPException(404)
    return FileResponse(row[0])


def get_setting(cur, key, default=None):
    cur.execute("SELECT value FROM app_settings WHERE key = %s", (key,))
    row = cur.fetchone()
    return row["value"] if row else default


def log_audit(cur, txn_id, action, username, before=None, after=None):
    cur.execute(
        """INSERT INTO audit_log (txn_id, action, changed_by, before_data, after_data)
           VALUES (%s, %s, %s, %s, %s)""",
        (
            txn_id,
            action,
            username,
            json.dumps(before, default=str, ensure_ascii=False) if before is not None else None,
            json.dumps(after, default=str, ensure_ascii=False) if after is not None else None,
        ),
    )


def save_uploaded_photo(cur, photo):
    if not photo or not photo.filename:
        return None
    ext = Path(photo.filename).suffix or ".jpg"
    local_path = LINE_FILES_DIR / f"{uuid.uuid4().hex}{ext}"
    local_path.write_bytes(photo.file.read())
    cur.execute(
        """INSERT INTO raw_files (file_type, storage_path, is_slip, processed)
           VALUES ('image', %s, false, true) RETURNING id""",
        (str(local_path),),
    )
    return cur.fetchone()[0]


@app.get("/")
def dashboard(request: Request):
    redirect = require_login(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """SELECT id, raw_file_id, txn_date, txn_time, direction, category, amount, bank,
                          sender_name, receiver_name, memo, ai_model, verified_bank, status,
                          qr_trans_ref, printed_ref
                   FROM slip_transactions
                   ORDER BY txn_date DESC, txn_time DESC, id DESC"""
            )
            rows = cur.fetchall()

            cur.execute("SELECT COALESCE(SUM(cost_usd), 0) AS total FROM ai_usage_log")
            ai_spent = float(cur.fetchone()["total"])
            ai_starting_balance = float(get_setting(cur, "ai_starting_balance_usd", "0"))
            categories = [c["name"] for c in get_categories(cur)]
    finally:
        conn.close()

    groups = defaultdict(list)
    for r in rows:
        if r["qr_trans_ref"]:
            groups[r["qr_trans_ref"]].append(r)
    dup_color = {}
    dup_extra_ids = set()
    palette = ["dup-1", "dup-2", "dup-3", "dup-4", "dup-5"]
    i = 0
    for ref, group in groups.items():
        if len(group) > 1:
            for r in group:
                dup_color[r["id"]] = palette[i % len(palette)]
            i += 1
            canonical = min(
                group, key=lambda r: (r["txn_date"] or date.min, r["txn_time"] or time.min, r["id"])
            )
            dup_extra_ids.update(r["id"] for r in group if r["id"] != canonical["id"])

    counted_rows = [r for r in rows if r["id"] not in dup_extra_ids]
    expense = sum(float(r["amount"]) for r in counted_rows if r["direction"] == "expense")
    income = sum(float(r["amount"]) for r in counted_rows if r["direction"] == "income")

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "user": request.session.get("user"),
            "role": request.session.get("role"),
            "rows": rows,
            "dup_color": dup_color,
            "expense": expense,
            "income": income,
            "net": income - expense,
            "count": len(counted_rows),
            "ai_balance": ai_starting_balance - ai_spent,
            "ai_starting_balance": ai_starting_balance,
            "categories": categories,
        },
    )


def _period_filter(granularity, year, month, day):
    if granularity == "day":
        return "txn_date = %s", (date(year, month, day),)
    if granularity == "month":
        return "EXTRACT(year FROM txn_date) = %s AND EXTRACT(month FROM txn_date) = %s", (year, month)
    return "EXTRACT(year FROM txn_date) = %s", (year,)


def fetch_report_data(cur, granularity, year, month, day):
    if granularity not in ("year", "month", "day"):
        granularity = "year"

    cur.execute(
        """SELECT DISTINCT EXTRACT(year FROM txn_date)::int AS y
           FROM slip_transactions
           WHERE txn_date IS NOT NULL
           ORDER BY 1 DESC"""
    )
    years = [r[0] for r in cur.fetchall()]
    if not year:
        year = years[0] if years else date.today().year

    months = []
    if granularity in ("month", "day"):
        cur.execute(
            """SELECT DISTINCT EXTRACT(month FROM txn_date)::int AS m
               FROM slip_transactions
               WHERE txn_date IS NOT NULL AND EXTRACT(year FROM txn_date) = %s
               ORDER BY 1 DESC""",
            (year,),
        )
        months = [r[0] for r in cur.fetchall()]
        if not month:
            month = months[0] if months else date.today().month

    days = []
    if granularity == "day":
        cur.execute(
            """SELECT DISTINCT EXTRACT(day FROM txn_date)::int AS d
               FROM slip_transactions
               WHERE txn_date IS NOT NULL AND EXTRACT(year FROM txn_date) = %s
                     AND EXTRACT(month FROM txn_date) = %s
               ORDER BY 1 DESC""",
            (year, month),
        )
        days = [r[0] for r in cur.fetchall()]
        if not day:
            day = days[0] if days else 1

    if granularity == "day":
        cur.execute(
            """SELECT to_char(txn_time, 'HH24:00') AS bucket,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'expense'), 0) AS expense,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'income'), 0) AS income
               FROM slip_transactions
               WHERE txn_date = %s AND direction IN ('expense', 'income')
               GROUP BY 1
               ORDER BY 1""",
            (date(year, month, day),),
        )
    elif granularity == "month":
        cur.execute(
            """SELECT to_char(txn_date, 'DD') AS bucket,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'expense'), 0) AS expense,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'income'), 0) AS income
               FROM slip_transactions
               WHERE txn_date IS NOT NULL AND direction IN ('expense', 'income')
                     AND EXTRACT(year FROM txn_date) = %s AND EXTRACT(month FROM txn_date) = %s
               GROUP BY 1
               ORDER BY 1""",
            (year, month),
        )
    else:
        cur.execute(
            """SELECT to_char(date_trunc('month', txn_date), 'YYYY-MM') AS bucket,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'expense'), 0) AS expense,
                      COALESCE(SUM(amount) FILTER (WHERE direction = 'income'), 0) AS income
               FROM slip_transactions
               WHERE txn_date IS NOT NULL AND direction IN ('expense', 'income')
                     AND EXTRACT(year FROM txn_date) = %s
               GROUP BY 1
               ORDER BY 1""",
            (year,),
        )
    bucket_rows = cur.fetchall()

    period_sql, period_params = _period_filter(granularity, year, month, day)

    cur.execute(
        f"""SELECT COALESCE(category, 'ไม่ระบุหมวด') AS category, SUM(amount) AS total
            FROM slip_transactions
            WHERE txn_date IS NOT NULL AND direction = 'expense' AND {period_sql}
            GROUP BY 1
            ORDER BY 2 DESC""",
        period_params,
    )
    category_rows = [{"name": r[0], "total": float(r[1])} for r in cur.fetchall()]

    cur.execute(
        f"""SELECT COALESCE(receiver_name, 'ไม่ระบุ') AS receiver, SUM(amount) AS total
            FROM slip_transactions
            WHERE txn_date IS NOT NULL AND direction = 'expense' AND {period_sql}
            GROUP BY 1
            ORDER BY 2 DESC
            LIMIT 5""",
        period_params,
    )
    top_payees = [{"name": r[0], "total": float(r[1])} for r in cur.fetchall()]

    bucket_chart = {
        "labels": [r[0] for r in bucket_rows],
        "expense": [float(r[1]) for r in bucket_rows],
        "income": [float(r[2]) for r in bucket_rows],
    }
    total_expense = sum(bucket_chart["expense"])
    total_income = sum(bucket_chart["income"])

    if granularity == "day":
        period_label = f"{day} {THAI_MONTHS.get(month, month)} {year}"
    elif granularity == "month":
        period_label = f"{THAI_MONTHS.get(month, month)} {year}"
    else:
        period_label = f"ปี {year}"

    return {
        "granularity": granularity, "years": years, "months": months, "days": days,
        "year": year, "month": month, "day": day, "period_label": period_label,
        "bucket_chart": bucket_chart, "category_rows": category_rows, "top_payees": top_payees,
        "total_expense": total_expense, "total_income": total_income, "net": total_income - total_expense,
    }


@app.get("/reports")
def reports(request: Request, granularity: str = "month", year: int = 0, month: int = 0, day: int = 0):
    redirect = require_login(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor() as cur:
            data = fetch_report_data(cur, granularity, year, month, day)
    finally:
        conn.close()

    category_chart = {
        "labels": [c["name"] for c in data["category_rows"]],
        "amounts": [c["total"] for c in data["category_rows"]],
    }

    return templates.TemplateResponse(
        request,
        "reports.html",
        {
            "user": request.session.get("user"),
            "role": request.session.get("role"),
            "monthly_chart": json.dumps(data["bucket_chart"]),
            "category_chart": json.dumps(category_chart),
            "total_expense": data["total_expense"],
            "total_income": data["total_income"],
            "net": data["net"],
            "top_payees": data["top_payees"],
            "years": data["years"],
            "months": data["months"],
            "days": data["days"],
            "selected_year": data["year"],
            "selected_month": data["month"],
            "selected_day": data["day"],
            "granularity": data["granularity"],
            "period_label": data["period_label"],
            "thai_months": THAI_MONTHS,
        },
    )


COLOR_EXPENSE = "C0392B"
COLOR_INCOME = "1E8449"


def _style_table_sheet(ws, df, table_name, currency_cols):
    """Turn a freshly-written df sheet into a banded Excel Table with sane column widths and number formats."""
    n_rows, n_cols = df.shape
    last_row = n_rows + 1
    ref = f"A1:{get_column_letter(n_cols)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)

    for col_name in currency_cols:
        col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
        for r in range(2, last_row + 1):
            ws[f"{col_letter}{r}"].number_format = "#,##0.00"

    for i, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 45)


@app.get("/reports/export/excel")
def reports_export_excel(request: Request, granularity: str = "year", year: int = 0, month: int = 0, day: int = 0):
    redirect = require_login(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor() as cur:
            data = fetch_report_data(cur, granularity, year, month, day)
    finally:
        conn.close()

    summary_df = pd.DataFrame(
        [
            ("ช่วงเวลา", data["period_label"]),
            ("รายจ่าย", data["total_expense"]),
            ("รายรับ", data["total_income"]),
            ("สุทธิ", data["net"]),
        ],
        columns=["รายการ", "ค่า"],
    )
    category_df = pd.DataFrame(
        [(c["name"], c["total"]) for c in data["category_rows"]], columns=["หมวด", "ยอด (บาท)"]
    )
    payee_df = pd.DataFrame(
        [(i + 1, p["name"], p["total"]) for i, p in enumerate(data["top_payees"])],
        columns=["อันดับ", "ผู้รับ", "ยอด (บาท)"],
    )
    bucket_df = pd.DataFrame(
        {
            "ช่วง": data["bucket_chart"]["labels"],
            "รายจ่าย": data["bucket_chart"]["expense"],
            "รายรับ": data["bucket_chart"]["income"],
        }
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="สรุป", index=False)
        category_df.to_excel(writer, sheet_name="ตามหมวด", index=False)
        payee_df.to_excel(writer, sheet_name="Top ผู้รับเงิน", index=False)
        bucket_df.to_excel(writer, sheet_name="กราฟ", index=False)

        _style_table_sheet(writer.sheets["สรุป"], summary_df, "tbl_summary", ["ค่า"])
        _style_table_sheet(writer.sheets["ตามหมวด"], category_df, "tbl_category", ["ยอด (บาท)"])
        _style_table_sheet(writer.sheets["Top ผู้รับเงิน"], payee_df, "tbl_payee", ["ยอด (บาท)"])
        _style_table_sheet(writer.sheets["กราฟ"], bucket_df, "tbl_bucket", ["รายจ่าย", "รายรับ"])

        if len(bucket_df):
            bar = BarChart()
            bar.type = "col"
            bar.title = "รายรับ-รายจ่ายตามช่วงเวลา"
            bar.y_axis.title = "บาท"
            bar.y_axis.numFmt = "#,##0"
            bar.width, bar.height = 24, 12
            bar.gapWidth = 50
            ws = writer.sheets["กราฟ"]
            n = len(bucket_df) + 1
            bar.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=n), titles_from_data=True)
            bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
            bar.series[0].graphicalProperties.solidFill = COLOR_EXPENSE
            bar.series[1].graphicalProperties.solidFill = COLOR_INCOME
            bar.dataLabels = DataLabelList()
            bar.dataLabels.showVal = True
            ws.add_chart(bar, f"{get_column_letter(bucket_df.shape[1] + 2)}2")

        if len(category_df):
            # category_rows already comes from SQL ORDER BY total DESC; horizontal bar
            # reads top-to-bottom in that same order once the category axis is reversed.
            cat_bar = BarChart()
            cat_bar.type = "bar"
            cat_bar.title = "รายจ่ายตามหมวด"
            cat_bar.y_axis.numFmt = "#,##0"
            cat_bar.x_axis.scaling.orientation = "maxMin"
            cat_bar.width, cat_bar.height = 22, max(10, 2 * len(category_df))
            cat_bar.gapWidth = 60
            cat_bar.legend = None
            ws2 = writer.sheets["ตามหมวด"]
            n2 = len(category_df) + 1
            cat_bar.add_data(Reference(ws2, min_col=2, min_row=1, max_row=n2), titles_from_data=True)
            cat_bar.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=n2))
            cat_bar.series[0].graphicalProperties.solidFill = COLOR_EXPENSE
            cat_bar.dataLabels = DataLabelList()
            cat_bar.dataLabels.showVal = True
            ws2.add_chart(cat_bar, f"{get_column_letter(category_df.shape[1] + 2)}2")
    buf.seek(0)

    filename = f"report_summary_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/reports/export/pdf")
def reports_export_pdf(request: Request, granularity: str = "year", year: int = 0, month: int = 0, day: int = 0):
    redirect = require_login(request)
    if redirect:
        return redirect

    try:
        from export_report import build_summary_pdf
    except RuntimeError as e:
        raise HTTPException(500, str(e))

    conn = db()
    try:
        with conn.cursor() as cur:
            data = fetch_report_data(cur, granularity, year, month, day)
    finally:
        conn.close()

    buf = io.BytesIO()
    build_summary_pdf(
        data["period_label"], data["total_expense"], data["total_income"], data["net"],
        data["category_rows"], data["top_payees"], data["bucket_chart"], buf,
    )
    buf.seek(0)

    filename = f"report_summary_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def build_export_filter(direction, date_from, date_to, bank, status, name, category):
    where = []
    params = []
    if category:
        where.append("category = %s")
        params.append(category)
    if direction:
        where.append("direction = ANY(%s)")
        params.append(direction.split(","))
    if date_from:
        where.append("txn_date >= %s")
        params.append(date_from)
    if date_to:
        where.append("txn_date <= %s")
        params.append(date_to)
    if bank:
        where.append("bank = %s")
        params.append(bank)
    if status == "verified":
        where.append("verified_bank = true")
    elif status == "unverified":
        where.append("verified_bank = false")
    elif status == "dup":
        where.append("""qr_trans_ref IN (
            SELECT qr_trans_ref FROM slip_transactions
            WHERE qr_trans_ref IS NOT NULL GROUP BY qr_trans_ref HAVING COUNT(*) > 1
        )""")
    if name:
        where.append("(sender_name ILIKE %s OR receiver_name ILIKE %s)")
        params.extend([f"%{name}%", f"%{name}%"])
    return where, params


def fetch_export_df(direction, date_from, date_to, bank, status, name, category, columns="full"):
    where, params = build_export_filter(direction, date_from, date_to, bank, status, name, category)
    if columns == "print":
        select = """
            txn_date AS "วันที่", txn_time AS "เวลา", bank AS "ธนาคาร",
            direction AS "ประเภท", category AS "หมวด", amount AS "ยอดเงิน",
            receiver_name AS "ผู้รับ", memo AS "รายละเอียด"
        """
    else:
        select = """
            txn_date AS "วันที่", txn_time AS "เวลา", bank AS "ธนาคาร",
            direction AS "ประเภท", category AS "หมวด", amount AS "ยอดเงิน", fee AS "ค่าธรรมเนียม",
            sender_name AS "ผู้โอน", receiver_name AS "ผู้รับ", memo AS "รายละเอียด",
            qr_trans_ref AS "เลขอ้างอิง (QR)", verified_bank AS "ยืนยันธนาคาร"
        """
    query = f"SELECT {select} FROM slip_transactions"
    if where:
        query += " WHERE " + " AND ".join(where)
    query += " ORDER BY txn_date, txn_time"

    conn = db()
    try:
        return pd.read_sql(query, conn, params=params)
    finally:
        conn.close()


@app.get("/export/excel")
def export_excel(
    request: Request,
    direction: str = "",
    date_from: str = "",
    date_to: str = "",
    bank: str = "",
    status: str = "",
    name: str = "",
    category: str = "",
):
    redirect = require_login(request)
    if redirect:
        return redirect

    df = fetch_export_df(direction, date_from, date_to, bank, status, name, category)

    summary = (
        df.groupby("ประเภท")["ยอดเงิน"]
        .sum()
        .reindex(["expense", "income", "unknown"])
        .fillna(0)
        .rename({"expense": "รายจ่าย", "income": "รายรับ", "unknown": "ไม่ระบุ"})
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="รายการ", index=False)
        summary.to_frame("รวม (บาท)").to_excel(writer, sheet_name="สรุป")
    buf.seek(0)

    filename = f"ledger_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/pdf")
def export_pdf(
    request: Request,
    direction: str = "",
    date_from: str = "",
    date_to: str = "",
    bank: str = "",
    status: str = "",
    name: str = "",
    category: str = "",
):
    redirect = require_login(request)
    if redirect:
        return redirect

    try:
        from export_report import build_pdf
    except RuntimeError as e:
        raise HTTPException(500, str(e))

    df = fetch_export_df(direction, date_from, date_to, bank, status, name, category, columns="print")

    buf = io.BytesIO()
    build_pdf(df, buf)
    buf.seek(0)

    filename = f"ledger_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/settings/ai-balance")
def update_ai_balance(request: Request, value: float = Form(...)):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute(
                """INSERT INTO app_settings (key, value) VALUES ('ai_starting_balance_usd', %s)
                   ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value""",
                (str(value),),
            )
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


@app.get("/settings/categories")
def categories_view(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor() as cur:
            categories = get_categories(cur)
    finally:
        conn.close()

    return templates.TemplateResponse(
        request,
        "settings_categories.html",
        {
            "user": request.session.get("user"), "role": request.session.get("role"),
            "categories": categories,
        },
    )


@app.post("/settings/categories/new")
def categories_new(request: Request, name: str = Form(...), keywords: str = Form("")):
    redirect = require_admin(request)
    if redirect:
        return redirect

    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM categories")
            next_order = cur.fetchone()[0]
            cur.execute(
                "INSERT INTO categories (name, keywords, sort_order) VALUES (%s, %s, %s)",
                (name.strip(), kw_list, next_order),
            )
    finally:
        conn.close()

    return RedirectResponse("/settings/categories", status_code=302)


@app.post("/settings/categories/{cat_id}/edit")
def categories_edit(cat_id: int, request: Request, name: str = Form(...), keywords: str = Form("")):
    redirect = require_admin(request)
    if redirect:
        return redirect

    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute(
                "UPDATE categories SET name = %s, keywords = %s WHERE id = %s",
                (name.strip(), kw_list, cat_id),
            )
    finally:
        conn.close()

    return RedirectResponse("/settings/categories", status_code=302)


@app.post("/settings/categories/{cat_id}/delete")
def categories_delete(cat_id: int, request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("DELETE FROM categories WHERE id = %s", (cat_id,))
    finally:
        conn.close()

    return RedirectResponse("/settings/categories", status_code=302)


@app.post("/settings/categories/{cat_id}/move")
def categories_move(cat_id: int, request: Request, direction: str = Form(...)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    if direction not in ("up", "down"):
        raise HTTPException(400)

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("SELECT id, sort_order FROM categories ORDER BY sort_order, id")
            rows = cur.fetchall()
            ids = [r[0] for r in rows]
            idx = ids.index(cat_id)
            swap_idx = idx - 1 if direction == "up" else idx + 1
            if 0 <= swap_idx < len(rows):
                a_id, a_order = rows[idx]
                b_id, b_order = rows[swap_idx]
                cur.execute("UPDATE categories SET sort_order = %s WHERE id = %s", (b_order, a_id))
                cur.execute("UPDATE categories SET sort_order = %s WHERE id = %s", (a_order, b_id))
    finally:
        conn.close()

    return RedirectResponse("/settings/categories", status_code=302)


def resolve_display_names(messages):
    names = {}
    if not LINE_ACCESS_TOKEN:
        return names
    config = Configuration(access_token=LINE_ACCESS_TOKEN)
    with ApiClient(config) as api_client:
        api = MessagingApi(api_client)
        for m in messages:
            user_id = m.get("line_user_id")
            group_id = m.get("line_group_id")
            if not user_id or not group_id or user_id in names:
                continue
            try:
                profile = api.get_group_member_profile(group_id, user_id)
                names[user_id] = profile.display_name
            except Exception:
                pass
    return names


AUDIT_FIELD_LABELS = {
    "bank": "ธนาคาร", "txn_date": "วันที่", "txn_time": "เวลา", "amount": "ยอดเงิน", "fee": "ค่าธรรมเนียม",
    "sender_name": "ผู้โอน", "receiver_name": "ผู้รับ", "memo": "รายละเอียด",
    "direction": "ประเภท", "category": "หมวด", "status": "สถานะ",
}
AUDIT_ACTION_LABELS = {"create": "เพิ่มรายการ", "update": "แก้ไขรายการ", "delete": "ลบรายการ"}


def _normalize_for_diff(key, value):
    if value is None:
        return None
    if key in ("amount", "fee"):
        try:
            return f"{float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)
    if key == "txn_time":
        return str(value)[:5]
    return str(value)


def summarize_audit_entry(action, before, after):
    if action == "create":
        return f"{(after or {}).get('memo') or '-'} ฿{(after or {}).get('amount', '-')}"
    if action == "delete":
        return f"{(before or {}).get('memo') or '-'} ฿{(before or {}).get('amount', '-')}"
    if not before or not after:
        return "-"
    changes = []
    for key, label in AUDIT_FIELD_LABELS.items():
        old_val, new_val = before.get(key), after.get(key)
        if _normalize_for_diff(key, old_val) != _normalize_for_diff(key, new_val):
            changes.append(f"{label}: {old_val} → {new_val}")
    return "; ".join(changes) if changes else "ไม่มีการเปลี่ยนแปลง"


@app.get("/audit")
def audit_view(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """SELECT id, txn_id, action, changed_by, changed_at, before_data, after_data
                   FROM audit_log
                   ORDER BY changed_at DESC
                   LIMIT 500"""
            )
            entries = cur.fetchall()
    finally:
        conn.close()

    for e in entries:
        e["action_label"] = AUDIT_ACTION_LABELS.get(e["action"], e["action"])
        e["summary"] = summarize_audit_entry(e["action"], e["before_data"], e["after_data"])

    return templates.TemplateResponse(
        request,
        "audit.html",
        {"user": request.session.get("user"), "role": request.session.get("role"), "entries": entries},
    )


@app.get("/chat")
def chat_view(request: Request):
    redirect = require_login(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """SELECT 'text' AS kind, id, line_user_id, line_group_id, text,
                          NULL::integer AS raw_file_id, received_at
                       FROM line_messages
                   UNION ALL
                   SELECT file_type AS kind, id, line_user_id, line_group_id, NULL AS text,
                          id AS raw_file_id, received_at
                       FROM raw_files
                   ORDER BY received_at DESC
                   LIMIT 500"""
            )
            messages = list(reversed(cur.fetchall()))
    finally:
        conn.close()

    names = resolve_display_names(messages)

    return templates.TemplateResponse(
        request,
        "chat.html",
        {
            "messages": messages,
            "names": names,
            "user": request.session.get("user"),
            "role": request.session.get("role"),
        },
    )


@app.get("/transactions/new")
def new_form(request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    empty = {
        "id": None, "raw_file_id": None, "bank": "", "txn_date": "", "txn_time": "",
        "amount": "", "fee": 0, "sender_name": "", "sender_account": "",
        "receiver_name": "", "receiver_account": "", "memo": "",
        "direction": "expense", "category": "", "status": "reviewed",
    }
    conn = db()
    try:
        with conn.cursor() as cur:
            categories = [c["name"] for c in get_categories(cur)]
    finally:
        conn.close()

    return templates.TemplateResponse(
        request,
        "edit_transaction.html",
        {
            "txn": empty, "form_action": "/transactions/new", "categories": categories,
            "user": request.session.get("user"), "role": request.session.get("role"),
        },
    )


@app.post("/transactions/new")
def new_submit(
    request: Request,
    bank: str = Form(""),
    txn_date: str = Form(...),
    txn_time: str = Form(...),
    amount: float = Form(...),
    fee: float = Form(0),
    sender_name: str = Form(""),
    sender_account: str = Form(""),
    receiver_name: str = Form(""),
    receiver_account: str = Form(""),
    memo: str = Form(""),
    direction: str = Form("expense"),
    category: str = Form(""),
    status: str = Form("reviewed"),
    photo: UploadFile = File(None),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            final_category = category or guess_category(memo, get_categories(cur))
            raw_file_id = save_uploaded_photo(cur, photo)
            cur.execute(
                """INSERT INTO slip_transactions
                       (raw_file_id, bank, txn_date, txn_time, amount, fee, sender_name, sender_account,
                        receiver_name, receiver_account, memo, direction, category, status, ai_model)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (raw_file_id, bank, txn_date, txn_time, amount, fee, sender_name, sender_account,
                 receiver_name, receiver_account, memo, direction, final_category, status, "manual"),
            )
            new_id = cur.fetchone()[0]
            log_audit(
                cur, new_id, "create", request.session.get("user"),
                after={
                    "bank": bank, "txn_date": txn_date, "txn_time": txn_time, "amount": amount, "fee": fee,
                    "sender_name": sender_name, "receiver_name": receiver_name, "memo": memo,
                    "direction": direction, "category": final_category, "status": status,
                },
            )
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


@app.get("/transactions/{txn_id}/edit")
def edit_form(txn_id: int, request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM slip_transactions WHERE id = %s", (txn_id,))
            txn = cur.fetchone()
            categories = get_categories(cur)
    finally:
        conn.close()

    if not txn:
        raise HTTPException(404)
    if not txn["category"]:
        txn["category"] = guess_category(txn["memo"], categories)
    return templates.TemplateResponse(
        request,
        "edit_transaction.html",
        {
            "txn": txn, "form_action": f"/transactions/{txn_id}/edit",
            "categories": [c["name"] for c in categories],
            "user": request.session.get("user"), "role": request.session.get("role"),
        },
    )


@app.post("/transactions/{txn_id}/edit")
def edit_submit(
    txn_id: int,
    request: Request,
    bank: str = Form(""),
    txn_date: str = Form(...),
    txn_time: str = Form(...),
    amount: float = Form(...),
    fee: float = Form(0),
    sender_name: str = Form(""),
    sender_account: str = Form(""),
    receiver_name: str = Form(""),
    receiver_account: str = Form(""),
    memo: str = Form(""),
    direction: str = Form("unknown"),
    category: str = Form(""),
    status: str = Form("reviewed"),
    photo: UploadFile = File(None),
):
    redirect = require_admin(request)
    if redirect:
        return redirect

    final_category = category or None
    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("SELECT * FROM slip_transactions WHERE id = %s", (txn_id,))
            columns = [c.name for c in cur.description]
            before_row = cur.fetchone()
            before = dict(zip(columns, before_row)) if before_row else None

            new_raw_file_id = save_uploaded_photo(cur, photo)
            raw_file_id = new_raw_file_id if new_raw_file_id is not None else (before or {}).get("raw_file_id")

            cur.execute(
                """UPDATE slip_transactions SET
                       raw_file_id=%s, bank=%s, txn_date=%s, txn_time=%s, amount=%s, fee=%s,
                       sender_name=%s, sender_account=%s, receiver_name=%s, receiver_account=%s,
                       memo=%s, direction=%s, category=%s, status=%s
                   WHERE id=%s""",
                (raw_file_id, bank, txn_date, txn_time, amount, fee, sender_name, sender_account,
                 receiver_name, receiver_account, memo, direction, final_category, status, txn_id),
            )
            log_audit(
                cur, txn_id, "update", request.session.get("user"),
                before=before,
                after={
                    "bank": bank, "txn_date": txn_date, "txn_time": txn_time, "amount": amount, "fee": fee,
                    "sender_name": sender_name, "receiver_name": receiver_name, "memo": memo,
                    "direction": direction, "category": final_category, "status": status,
                },
            )
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


@app.post("/transactions/{txn_id}/direction")
def update_direction(txn_id: int, request: Request, direction: str = Form(...)):
    redirect = require_admin(request)
    if redirect:
        return redirect
    if direction not in ("expense", "income", "unknown"):
        raise HTTPException(400)

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("SELECT direction FROM slip_transactions WHERE id = %s", (txn_id,))
            old = cur.fetchone()
            cur.execute("UPDATE slip_transactions SET direction = %s WHERE id = %s", (direction, txn_id))
            log_audit(
                cur, txn_id, "update", request.session.get("user"),
                before={"direction": old[0] if old else None}, after={"direction": direction},
            )
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


@app.post("/transactions/{txn_id}/category")
def update_category(txn_id: int, request: Request, category: str = Form("")):
    redirect = require_admin(request)
    if redirect:
        return redirect

    final_category = category or None
    conn = db()
    try:
        with conn, conn.cursor() as cur:
            if category:
                valid_names = [c["name"] for c in get_categories(cur)]
                if category not in valid_names:
                    raise HTTPException(400)
            cur.execute("SELECT category FROM slip_transactions WHERE id = %s", (txn_id,))
            old = cur.fetchone()
            cur.execute("UPDATE slip_transactions SET category = %s WHERE id = %s", (final_category, txn_id))
            log_audit(
                cur, txn_id, "update", request.session.get("user"),
                before={"category": old[0] if old else None}, after={"category": final_category},
            )
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


@app.post("/transactions/{txn_id}/delete")
def delete_submit(txn_id: int, request: Request):
    redirect = require_admin(request)
    if redirect:
        return redirect

    conn = db()
    try:
        with conn, conn.cursor() as cur:
            cur.execute("SELECT * FROM slip_transactions WHERE id = %s", (txn_id,))
            columns = [c.name for c in cur.description]
            before_row = cur.fetchone()
            before = dict(zip(columns, before_row)) if before_row else None

            cur.execute("DELETE FROM slip_transactions WHERE id = %s", (txn_id,))
            log_audit(cur, txn_id, "delete", request.session.get("user"), before=before)
    finally:
        conn.close()

    return RedirectResponse("/", status_code=302)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8081)
